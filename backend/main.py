from typing import Optional
from fastapi import FastAPI
import pymongo
import os
import shutil
from db import DB
from bson.json_util import dumps
from bson import ObjectId
import json
from fastapi.middleware.cors import CORSMiddleware
from collections import Counter
from fastapi import FastAPI, File, UploadFile
from openpyxl import load_workbook



app = FastAPI()

origins = [
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

database = DB().connect()
# connect to database
db = database.test
# connect to collection
collection = db.test

def bson_to_json(data):
    return json.loads(dumps(data))


@app.get("/cars/")
def cars(q: Optional[str] = None):
    if q is not None:
        cars = bson_to_json(collection.find({"manufacturer_name": q}))
    else:
        cars = bson_to_json(collection.find({}))
    return cars


@app.get("/car/{item_id}")
def car(item_id: str):
    car = collection.find_one({'_id': ObjectId(item_id)})
    car = bson_to_json(car)
    return car


@app.post("/car/add/")
def car_add(data: dict):
    car = collection.insert_one(data)
    car = bson_to_json(collection.find_one({'_id': ObjectId(car.inserted_id)}))
    return car


@app.patch("/car/update/{item_id}")
def car_update(item_id: str, data: dict):
    car = collection.find_one_and_update(
        {'_id': ObjectId(item_id)},  
        { "$set": data }, 
        return_document=pymongo.ReturnDocument.AFTER
    )
    car = bson_to_json(car)
    return car


@app.delete("/car/delete/{item_id}")
def car_delete(item_id: str):
    car = collection.find_one_and_delete({'_id': ObjectId(item_id)})
    text = "{} has been deleted successfully".format(item_id)
    return {"details": text}


@app.get("/car/search/")
def car_search(q: Optional[str] = None):
    # cars = collection.create_index([('name', 'text'), ('manufacturer_name', 'text')])
    search_term = "^{}".format(q)
    cars = collection.find({'manufacturer_name': {'$regex': search_term, "$options" :'i'}}, {"_id": 0, "manufacturer_name": 1} ).distinct("manufacturer_name")
    # cars = collection.find({ "$or": [{"name": q}]} )
    cars = bson_to_json(cars)
    return cars


@app.post("/car/bulk/add/")
def car_bulk_add(data: list):
    car = collection.insert_many(data)
    text = {"details": "Documents inserted succesfully"}
    return text


@app.get("/details/")
def get_details(q: Optional[str] = None):
    if q is not None:
        data = bson_to_json(collection.find({"manufacturer_name": q}, {"_id": 0, "manufacturer_name": 1}))
    else:
        data = bson_to_json(collection.find({}, {"_id": 0, "manufacturer_name": 1}))
    manufacturer_name = Counter([list(i.values())[0] for i in data if list(i.values())!=[]])
    return manufacturer_name

def excel_to_json(data_path):
    workbook = load_workbook(data_path)
    sheet = workbook.active
    products = []
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        product = {
            "manufacturer_name": row[0],
            "model": row[1],
            "year": row[2],
            "producing_country_name": row[3]
        }
        products.append(product)
    print(products)
    collection.insert_many(products)
    return None

@app.post("/uploadfile/")
def create_upload_file(file: UploadFile = File(...)):
    location="{}/media/{}".format(os.getcwd(), file.filename)    
    # create empty file to copy the file_object to
    upload_folder = open(location, 'wb+')
    shutil.copyfileobj(file.file, upload_folder )
    # upload_folder.close()
    excel_to_json(location)
    return {"filename": file.filename}

    